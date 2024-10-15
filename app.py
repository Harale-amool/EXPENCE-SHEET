from fastapi import FastAPI, Form, Request
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# Initialize FastAPI app
app = FastAPI()

# Set up Jinja2 templates
templates = Jinja2Templates(directory="templates")

# Define the path for the Excel file
excel_file = 'expenses.xlsx'

# Create Excel file with headers if it doesn't exist
def create_excel_file():
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Item", "Amount", "Person1", "Person2"])  # Headers
        wb.save(excel_file)

create_excel_file()

@app.get("/", response_class=HTMLResponse)
async def form(request: Request):
    # Render the form page
    return templates.TemplateResponse("form.html", {"request": request})

@app.post("/submit", response_class=RedirectResponse)
async def submit(request: Request, item: str = Form(...), amount: float = Form(...)):
    # Divide the amount between two people
    person1_share = amount / 2
    person2_share = amount / 2
    
    # Get current date
    current_date = datetime.now().strftime('%Y-%m-%d')
    
    # Load the workbook and append the new data
    wb = load_workbook(excel_file)
    ws = wb.active
    ws.append([current_date, item, amount, person1_share, person2_share])
    wb.save(excel_file)
    
    # Redirect to the data display page
    return RedirectResponse(url="/show", status_code=303)

@app.get("/show", response_class=HTMLResponse)
async def show(request: Request):
    # Read the Excel file content
    wb = load_workbook(excel_file)
    ws = wb.active
    
    # Convert the Excel data to a list of rows (for the table)
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)
    
    # Pass the data to the template to render it as a table
    return templates.TemplateResponse("show_data.html", {"request": request, "data": data})

